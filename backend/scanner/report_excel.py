"""
Export Excel (.xlsx) des rapports CyberScan via openpyxl.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Scan
from .report_generator import (
    _collect_recommendations,
    _collect_tools_used,
    build_report_context,
    risk_level_from_score,
    security_score_from_risk,
)

HEADER_FILL = PatternFill('solid', fgColor='0F172A')
HEADER_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _auto_width(ws, max_col: int = 6):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max_len + 2


def _write_header_row(ws, headers: List[str], row: int = 1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _technologies_rows(resultats: dict) -> List[List[Any]]:
    ww = resultats.get('whatweb') or {}
    techs = ww.get('technologies') or []
    rows = []
    for tech in techs:
        rows.append([
            tech.get('name', ''),
            ', '.join(str(v) for v in (tech.get('version') or [])),
            ', '.join(str(s) for s in (tech.get('string') or [])),
        ])
    return rows


def _vulnerabilities_rows(resultats: dict) -> List[List[Any]]:
    rows = []
    for vuln in resultats.get('vulnerabilities') or []:
        rows.append([str(vuln), 'SSL/TLS', 'Détecté par SSLScan'])

    for finding in resultats.get('zap_findings') or []:
        rows.append([
            finding.get('name', ''),
            finding.get('risk', ''),
            (finding.get('description') or '')[:500],
        ])

    for finding in resultats.get('nuclei_findings') or []:
        rows.append([
            finding.get('name', ''),
            finding.get('severity', ''),
            (finding.get('description') or '')[:500],
        ])
    return rows


def build_excel_workbook(scan: Scan) -> Workbook:
    context = build_report_context(scan)
    resultats = scan.resultats_ssl or {}
    risk_label, _ = risk_level_from_score(scan.score_risque_ia)
    security_score = security_score_from_risk(scan.score_risque_ia)
    tools = _collect_tools_used(resultats)
    recommendations = _collect_recommendations(scan, resultats)

    wb = Workbook()

    # --- Informations générales ---
    ws_info = wb.active
    ws_info.title = 'Informations générales'
    info_rows = [
        ('Domaine', scan.domaine),
        ('Date du scan', scan.date_scan.strftime('%d/%m/%Y %H:%M') if scan.date_scan else '—'),
        ('Score IA (risque)', f'{scan.score_risque_ia}/10'),
        ('Score global sécurité', f'{security_score}/10'),
        ('Niveau de risque', risk_label),
        ('Outils utilisés', ', '.join(tools)),
        ('Nombre de CVE', len(context.get('cves') or [])),
        ('Nombre de vulnérabilités', len(resultats.get('vulnerabilities') or [])),
    ]
    ws_info['A1'] = 'Champ'
    ws_info['B1'] = 'Valeur'
    for col, header in enumerate(['Champ', 'Valeur'], 1):
        cell = ws_info.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for idx, (key, value) in enumerate(info_rows, 2):
        ws_info.cell(row=idx, column=1, value=key).font = Font(bold=True)
        ws_info.cell(row=idx, column=2, value=value)
    _auto_width(ws_info, 2)

    # --- Technologies ---
    ws_tech = wb.create_sheet('Technologies détectées')
    _write_header_row(ws_tech, ['Technologie', 'Version', 'Détails'])
    tech_rows = _technologies_rows(resultats)
    for row_idx, row in enumerate(tech_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws_tech.cell(row=row_idx, column=col_idx, value=value)
    if not tech_rows:
        ws_tech.cell(row=2, column=1, value='Aucune technologie détectée')
    _auto_width(ws_tech, 3)

    # --- Vulnérabilités ---
    ws_vuln = wb.create_sheet('Vulnérabilités')
    _write_header_row(ws_vuln, ['Vulnérabilité', 'Sévérité', 'Description'])
    vuln_rows = _vulnerabilities_rows(resultats)
    for row_idx, row in enumerate(vuln_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws_vuln.cell(row=row_idx, column=col_idx, value=value)
    if not vuln_rows:
        ws_vuln.cell(row=2, column=1, value='Aucune vulnérabilité détectée')
    _auto_width(ws_vuln, 3)

    # --- CVE ---
    ws_cve = wb.create_sheet('CVE')
    _write_header_row(ws_cve, ['CVE ID', 'CVSS', 'Produit concerné', 'Description', 'Lien NVD', 'Recommandation IA'])
    for row_idx, cve in enumerate(context.get('cves') or [], 2):
        ws_cve.cell(row=row_idx, column=1, value=cve.get('cve_id'))
        ws_cve.cell(row=row_idx, column=2, value=cve.get('cvss_score'))
        ws_cve.cell(row=row_idx, column=3, value=cve.get('produit_concerne'))
        ws_cve.cell(row=row_idx, column=4, value=(cve.get('description') or '')[:500])
        ws_cve.cell(row=row_idx, column=5, value=cve.get('lien_nvd'))
        ws_cve.cell(row=row_idx, column=6, value=(cve.get('recommandation_ia') or '')[:500])
    if not context.get('cves'):
        ws_cve.cell(row=2, column=1, value='Aucune CVE enregistrée')
    _auto_width(ws_cve, 6)

    # --- Recommandations ---
    ws_rec = wb.create_sheet('Recommandations')
    _write_header_row(ws_rec, ['#', 'Recommandation'])
    for row_idx, rec in enumerate(recommendations, 2):
        ws_rec.cell(row=row_idx, column=1, value=row_idx - 1)
        ws_rec.cell(row=row_idx, column=2, value=rec)
    if not recommendations:
        ws_rec.cell(row=2, column=2, value='Aucune recommandation')
    _auto_width(ws_rec, 2)

    return wb


def generate_excel_bytes(scan: Scan) -> bytes:
    wb = build_excel_workbook(scan)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def excel_filename_for_scan(scan: Scan) -> str:
    safe_domain = re.sub(r'[^\w.\-]+', '_', scan.domaine)[:80]
    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return f'rapport_cyberscan_{scan.id}_{safe_domain}_{stamp}.xlsx'
